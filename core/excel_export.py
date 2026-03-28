import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime, date

class ExcelExporter:
    def __init__(self, title, company_name, filters=None):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Report"
        self.title = title
        self.company_name = company_name
        self.filters = filters or {}
        
        # Styles
        self.header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.title_font = Font(size=16, bold=True)
        self.bold_font = Font(bold=True)
        self.border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
        self.alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        self.red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    def _apply_header(self, columns):
        # Company Name
        self.ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        cell = self.ws.cell(row=1, column=1, value=self.company_name)
        cell.font = self.title_font
        cell.alignment = Alignment(horizontal="center")

        # Report Title
        self.ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
        cell = self.ws.cell(row=2, column=1, value=self.title)
        cell.font = self.bold_font
        cell.alignment = Alignment(horizontal="center")

        # Filters
        filter_str = " | ".join([f"{k}: {v}" for k, v in self.filters.items()])
        self.ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(columns))
        cell = self.ws.cell(row=3, column=1, value=filter_str)
        cell.alignment = Alignment(horizontal="center")

        # Column Headers
        for i, col in enumerate(columns, 1):
            cell = self.ws.cell(row=5, column=i, value=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center")
        
        self.ws.freeze_panes = "A6"

    def export(self, columns, data, highlight_overdue=False):
        self._apply_header(columns)
        
        row_num = 6
        for row_data in data:
            for col_num, value in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_num, column=col_num, value=value)
                
                # Alternating Row Color
                if row_num % 2 != 0:
                    cell.fill = self.alt_fill
                
                # Formatting based on type
                if isinstance(value, (int, float, complex)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                elif isinstance(value, (datetime, date)):
                    cell.number_format = 'DD-MM-YYYY'
                    cell.alignment = Alignment(horizontal="center")
                
                # Overdue Highlighting (assumes last column or specific logic)
                if highlight_overdue and "overdue" in str(row_data).lower(): # Simple heuristic
                     cell.fill = self.red_fill

            row_num += 1

        # Auto-adjust column width
        for col_index, col_cells in enumerate(self.ws.columns, 1):
            max_length = 0
            column = get_column_letter(col_index)
            for cell in col_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            self.ws.column_dimensions[column].width = adjusted_width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{self.title}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        self.wb.save(response)
        return response
